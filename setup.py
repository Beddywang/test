# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font,colors
from openpyxl.styles import NamedStyle, Font, Border, Side,PatternFill
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = Workbook()
ws = wb.active

ft = Font(name=u'微软雅黑',
    size=11,
    bold=False,
    italic=False,
    vertAlign=None,
    underline='none',
    strike=False,
    color='FF100002')

fill = PatternFill(fill_type="solid",
    start_color='FFEEFFFF',
    end_color='FF001100')

#边框可以选择的值为：'hair', 'medium', 'dashDot', 'dotted', 'mediumDashDot', 'dashed', 'mediumDashed', 'mediumDashDotDot', 'dashDotDot', 'slantDashDot', 'double', 'thick', 'thin']
#diagonal 表示对角线
bd = Border(left=Side(border_style="thin",
              color='FF001000'),
    right=Side(border_style="thin",
               color='FF110000'),
    top=Side(border_style="thin",
             color='FF110000'),
    bottom=Side(border_style="thin",
                color='FF110000'),
    diagonal=Side(border_style=None,
                  color='FF000000'),
    diagonal_direction=0,
    outline=Side(border_style=None,
                 color='FF000000'),
    vertical=Side(border_style=None,
                  color='FF000000'),
    horizontal=Side(border_style=None,
                   color='FF120000')
                )

alignment=Alignment(horizontal='general',
        vertical='bottom',
        text_rotation=0,
        wrap_text=False,
        shrink_to_fit=False,
        indent=0)

number_format = 'General'

protection = Protection(locked=True,
            hidden=False)

ws["B5"].font = ft
ws["B5"].fill =fill
ws["B5"].border = bd
ws["B5"].alignment = alignment
ws["B5"].number_format = number_format

ws["B5"].value ="glory road"
# ws["B5"].font = Font(color=colors.RED)
ws["B5"].font = Font(color=colors.RED, italic=True)

# Save the file
wb.save("e:\\sample.xlsx")
