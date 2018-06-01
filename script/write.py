# -*- coding: utf-8 -*-
import json
from openpyxl import Workbook
import datetime
from openpyxl import load_workbook


def write_json(path, version, imagename, count, appcounts):
    wb = load_workbook(path)
    wb.guess_types = True
    ws1 = wb.active
    # ws1 = wb.create_sheet(u"blade测试结果",0)
    list_title = [u"测试时间", u"测试版本", u"测试镜像名称", u"提取结果"]
    try:
        for i in appcounts:
            list_title.append(i[0])
        ws1.append(list_title)
        list_result = [datetime.datetime.now(),version,imagename,count]
        for i in appcounts:
            list_result.append(i[1])
        ws1.append(list_result)
    except Exception as e:
        print e
    wb.save(path)




if __name__ == '__main__':

    write_json("E:\\python_file\\Exercise\\blade-test\\blade-test\\source\\source.xlsx","1.0",u'测试.dmg',126,[(u'reg', u'0'), (u'usb', u'0')])